"""One-off: Forecast PnL workbook → server/data/workbook.json"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

SRC = Path("/Users/zainulmistry/Downloads/Switch Rentals/Forecast PnL from April 2026 to March 2027.xlsx")
OUT = Path(__file__).resolve().parents[1] / "server/data/workbook.json"

FY = [f"2026-{m:02d}" for m in range(4, 13)] + [f"2027-{m:02d}" for m in range(1, 4)]

SHEET_META = {
    "capex": {"id": "capex", "label": "CapEx", "group": "capex"},
    "Pnl": {"id": "pnl", "label": "P&L", "group": None},
    "Refund to Samir for all purchas": {"id": "refund-samir", "label": "Refund to Samir", "group": "capex"},
    "Cash flow ": {"id": "cash-flow", "label": "Cash flow", "group": None},
    "Refund to clients ": {"id": "refund-clients", "label": "Refund to clients", "group": "admin"},
    "Rent ": {"id": "rent", "label": "Rent", "group": "fixed"},
    "Salary ": {"id": "salary", "label": "Salary", "group": "fixed"},
    "Eletricity  & Water  ": {"id": "utilities", "label": "Electricity & water", "group": "fixed"},
    "CA ,Client refun D, Fire ,BMC ,": {"id": "ca-bmc", "label": "CA, Fire, BMC, GST", "group": "admin"},
    "OPS Supply": {"id": "ops-supply", "label": "OPS supply", "group": "ops"},
    "Business Dev F&B ": {"id": "bd-fnb", "label": "Business development / F&B", "group": "admin"},
    "Transport ": {"id": "transport", "label": "Transport", "group": "ops"},
    "New purchase CCG": {"id": "new-ccg", "label": "New purchase CCG", "group": "capex"},
    "Petty cash ": {"id": "petty", "label": "Petty cash", "group": "fixed"},
    "internet  ": {"id": "internet", "label": "Internet", "group": "fixed"},
    "AMC": {"id": "amc", "label": "AMC", "group": "nonctrl"},
}

CATEGORY_BY_SHEET = {
    "refund-samir": "Refund to Samir",
    "refund-clients": "Client refunds",
    "rent": "Rent with TDS",
    "salary": "Salary",
    "utilities": "Electricity & water",
    "ca-bmc": "CA, Fire, BMC, GST filing",
    "ops-supply": "OPS supply",
    "bd-fnb": "Business development",
    "transport": "Transport",
    "new-ccg": "New CCG purchase",
    "petty": "Petty cash",
    "internet": "Internet",
    "amc": "AMC",
}


def cell(v):
    if v is None or v == "":
        return None
    if isinstance(v, float):
        return round(v, 2)
    if isinstance(v, int):
        return v
    return str(v).strip()


def money(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return 0.0


def skip_label(label) -> bool:
    if label is None:
        return True
    s = str(label).strip().lower()
    if not s:
        return True
    if s in {"year", "cost"}:
        return True
    if s.startswith("total"):
        return True
    if s.startswith("balance"):
        return True
    if "total payout" in s:
        return True
    if "on 1st of the month" in s:
        return True
    return False


def last_day(month: str) -> str:
    from calendar import monthrange

    y, m = map(int, month.split("-"))
    return f"{y}-{m:02d}-{monthrange(y, m)[1]:02d}"


def grid(ws):
    rows = []
    for r in range(1, (ws.max_row or 1) + 1):
        row = [cell(ws.cell(r, c).value) for c in range(1, (ws.max_column or 1) + 1)]
        if any(x is not None and x != "" for x in row):
            rows.append(row)
    return rows


def months_from_header_row(row) -> list[tuple[int, str]]:
    """Return (col_index_0based, month_key) for a header row."""
    found = []
    text = " ".join(str(x or "") for x in row).lower()
    if "april and may" in text or "march" in text and "april and may" in text:
        # Salary / electricity: B March (skip), C April+May, D June...
        # col0 label, col1 March 2026 skip, col2 combined Apr/May
        mapping = [
            (2, "2026-04"),
            (2, "2026-05"),  # same column, split later
            (3, "2026-06"),
            (4, "2026-07"),
            (5, "2026-08"),
            (6, "2026-09"),
            (7, "2026-10"),
            (8, "2026-11"),
            (9, "2026-12"),
            (10, "2027-01"),
            (11, "2027-02"),
            (12, "2027-03"),
        ]
        return mapping
    # Standard: B Apr ... M Mar
    for i, mk in enumerate(FY):
        found.append((i + 1, mk))
    return found


def extract_monthly_lines(ws, sheet_id: str, group: str, category: str, split_combined: bool):
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    header_idx = 1 if len(rows) > 1 else 0
    header = [cell(x) for x in (rows[header_idx] if rows else [])]
    mapping = months_from_header_row(header)
    combined_col = 2 if split_combined else None
    lines = []
    start = header_idx + 1
    for raw in rows[start:]:
        label = cell(raw[0]) if raw else None
        if skip_label(label):
            continue
        seen_combined = False
        for col_i, mk in mapping:
            if col_i >= len(raw):
                continue
            amt = money(raw[col_i])
            if abs(amt) < 0.005:
                continue
            if split_combined and col_i == combined_col:
                if seen_combined:
                    continue
                seen_combined = True
                half = round(amt / 2, 2)
                for month in ("2026-04", "2026-05"):
                    lines.append(
                        {
                            "sheet": sheet_id,
                            "month": month,
                            "spentOn": last_day(month),
                            "costGroup": group,
                            "category": category,
                            "description": f"{label} (April & May split)",
                            "vendor": str(label),
                            "amount": half,
                        }
                    )
                continue
            lines.append(
                {
                    "sheet": sheet_id,
                    "month": mk,
                    "spentOn": last_day(mk),
                    "costGroup": group,
                    "category": category,
                    "description": str(label),
                    "vendor": str(label),
                    "amount": round(amt, 2),
                }
            )
    return lines


def main():
    wb = load_workbook(SRC, data_only=True)
    sheets = []
    expenses = []
    cash = []
    capital = []
    capex_opening = []

    for name in wb.sheetnames:
        ws = wb[name]
        meta = SHEET_META.get(name) or {"id": name.strip().lower(), "label": name.strip(), "group": "admin"}
        g = grid(ws)
        sheets.append({"id": meta["id"], "excelName": name, "label": meta["label"], "group": meta["group"], "grid": g})

        sid = meta["id"]
        if sid in {"pnl", "cash-flow", "capex", "new-ccg"}:
            continue
        split = sid in {"salary", "utilities"}
        expenses.extend(
            extract_monthly_lines(ws, sid, meta["group"] or "admin", CATEGORY_BY_SHEET[sid], split)
        )

        if sid == "refund-samir":
            # Company owes Samir for personal buys — also a partner draw (cash out to him).
            by_month = {}
            for line in expenses:
                if line["sheet"] == "refund-samir":
                    by_month[line["month"]] = by_month.get(line["month"], 0) + line["amount"]
            for month, amt in by_month.items():
                capital.append(
                    {
                        "partner": "samir",
                        "kind": "draw",
                        "amount": round(amt, 2),
                        "occurredOn": last_day(month),
                        "notes": "Refund to Samir for purchases on his account (workbook)",
                    }
                )

    # CapEx sheet: historical mill setup + old CCG vendors (already paid).
    cap = wb["capex"]
    # FY crockery buys from Cash flow "Payment towards Asset" (full vendor list).
    cf = wb["Cash flow "]
    for r in range(45, 56):
        vendor = cell(cf.cell(r, 1).value)
        if skip_label(vendor) or vendor is None:
            continue
        if str(vendor).lower().startswith("total"):
            continue
        for i, mk in enumerate(FY):
            amt = money(cf.cell(r, i + 2).value)
            if abs(amt) < 0.005:
                continue
            expenses.append(
                {
                    "sheet": "new-ccg",
                    "month": mk,
                    "spentOn": last_day(mk),
                    "costGroup": "capex",
                    "category": "New CCG purchase",
                    "description": str(vendor),
                    "vendor": str(vendor),
                    "amount": round(amt, 2),
                }
            )

    # Opening CCG already on the mill (capex sheet), not this year's Amazon/Harbour buys.
    opening_ccg = money(cap.cell(13, 2).value)
    if opening_ccg:
        capex_opening.append(
            {
                "sheet": "capex",
                "month": "2026-04",
                "spentOn": "2026-04-01",
                "costGroup": "capex",
                "category": "Opening CCG stock",
                "description": "CCG purchased before / at go-live (Harbour, Thanor, Nirman, Manekia, Neeti)",
                "vendor": "Opening CCG",
                "amount": round(opening_ccg, 2),
            }
        )
    vendors = [
        (3, "Harbour Artevo, Selite, Rona"),
        (4, "Thanor Pottery"),
        (5, "Viraj Nirman"),
        (6, "SK Manekia"),
        (7, "SK Manekia"),
        (8, "Neeti Udyog (old)"),
    ]
    # Opening CCG total is already in row 13; vendor split is memo only (do not double-count).
    capex_vendors = []
    for row, vendor in vendors:
        amt = money(cap.cell(row, 9).value)
        if amt:
            capex_vendors.append({"vendor": vendor, "amount": round(amt, 2), "paid": round(money(cap.cell(row, 10).value), 2)})

    # Cash flow: actual bank / cash by month (row 18 / 19, cols B-M = Apr-Mar)
    cf = wb["Cash flow "]
    for i, mk in enumerate(FY):
        bank = money(cf.cell(18, i + 2).value)
        petty = money(cf.cell(19, i + 2).value)
        if bank or petty:
            cash.append({"asOf": last_day(mk), "bankAmount": round(bank, 2), "cashAmount": round(petty, 2), "notes": "From Cash flow tab (actual in bank)"})

    # Do not book refund-to-Samir as capex AND draw AND CCG — keep as capex reimbursement only if not already on New CCG.
    # Amazon on refund-samir is reimbursement; New purchase CCG already has Amazon. Drop refund-samir from expenses (keep draw).
    expenses = [e for e in expenses if e["sheet"] != "refund-samir"]

    # Skip prior-year March leftover if any leaked
    expenses = [e for e in expenses if e["month"] in FY]

    payload = {
        "source": str(SRC.name),
        "sheets": sheets,
        "expenses": expenses + capex_opening,
        "cash": cash,
        "capitalDraws": capital,
        "capexVendors": capex_vendors,
        "partners": {
            "samirTarget": 4810000,
            "karanTarget": 1690000,
            "samirInvested": 4810000,
            "karanInvested": 1690000,
        },
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    by_sheet = {}
    for e in payload["expenses"]:
        by_sheet[e["sheet"]] = by_sheet.get(e["sheet"], 0) + e["amount"]
    print("wrote", OUT)
    print("expense lines", len(payload["expenses"]), "total", round(sum(e["amount"] for e in payload["expenses"]), 2))
    print("by sheet", {k: round(v, 2) for k, v in sorted(by_sheet.items())})
    print("cash months", len(cash), "draws", capital)


if __name__ == "__main__":
    main()
