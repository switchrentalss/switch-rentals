#!/usr/bin/env python3
"""Import Switch Rental live Excel workbooks into server/data/switch-finance.json."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

FOLDER = Path("/Users/zainulmistry/Downloads/Switch Rentals")
OUT = Path(__file__).resolve().parents[1] / "server" / "data" / "switch-finance.json"

MONTHS = [
    ("2026-04", "Apr 2026", "April 26"),
    ("2026-05", "May 2026", "May 26"),
    ("2026-06", "Jun 2026", "June 2026"),
    ("2026-07", "Jul 2026", "July 2026"),
    ("2026-08", "Aug 2026", "Aug 2026"),
    ("2026-09", "Sep 2026", "Sep 2026"),
]

BUDGET_SHARE = {
    "2026-04": 0.07,
    "2026-05": 0.07,
    "2026-06": 0.04,
    "2026-07": 0.04,
    "2026-08": 0.04,
    "2026-09": 0.06,
    "2026-10": 0.10,
    "2026-11": 0.12,
    "2026-12": 0.12,
    "2027-01": 0.12,
    "2027-02": 0.12,
    "2027-03": 0.10,
}

FY_MONTHS = list(BUDGET_SHARE.keys())
ANNUAL_BUDGET = 7_000_000


def to_num(value) -> float:
    if value is None or value is False:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("\xa0", "")
    if text in ("", "-", "None"):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


class WorkbookEngine:
    def __init__(self, path: Path):
        self.wb = load_workbook(path, data_only=False)
        self.cache: dict[tuple[str, int, int], float | None] = {}
        self.visiting: set[tuple[str, int, int]] = set()

    def sheet(self, name: str):
        if name in self.wb.sheetnames:
            return self.wb[name]
        stripped = name.strip()
        for existing in self.wb.sheetnames:
            if existing.strip() == stripped:
                return self.wb[existing]
        raise KeyError(name)

    def cell(self, sheet: str, row: int, col: int) -> float:
        key = (sheet, row, col)
        if key in self.cache:
            return self.cache[key] or 0.0
        if key in self.visiting:
            return 0.0
        self.visiting.add(key)
        ws = self.sheet(sheet)
        raw = ws.cell(row, col).value
        result = self.eval_value(raw, sheet)
        self.cache[key] = result
        self.visiting.remove(key)
        return result

    def eval_value(self, raw, current_sheet: str) -> float:
        if raw is None:
            return 0.0
        if isinstance(raw, (int, float)):
            return float(raw)
        if not isinstance(raw, str):
            return 0.0
        text = raw.strip()
        if not text.startswith("="):
            return to_num(text)
        return self.eval_formula(text[1:], current_sheet)

    def eval_formula(self, formula: str, current_sheet: str) -> float:
        formula = formula.replace("\n", " ").strip()
        formula = re.sub(r"\s+", "", formula)
        formula = formula.replace("%", "/100")

        def sheet_ref(match: re.Match) -> str:
            sheet = match.group(1).replace("''", "'")
            col = column_index_from_string(match.group(2))
            row = int(match.group(3))
            try:
                return str(self.cell(sheet, row, col))
            except KeyError:
                return "0"

        def sum_fn(match: re.Match) -> str:
            start, end = match.group(1), match.group(2)
            sc, sr = re.match(r"([A-Z]+)(\d+)", start).groups()
            ec, er = re.match(r"([A-Z]+)(\d+)", end).groups()
            c1, c2 = column_index_from_string(sc), column_index_from_string(ec)
            r1, r2 = int(sr), int(er)
            total = 0.0
            for r in range(min(r1, r2), max(r1, r2) + 1):
                for c in range(min(c1, c2), max(c1, c2) + 1):
                    total += self.cell(current_sheet, r, c)
            return str(total)

        while True:
            nxt = re.sub(r"SUM\(([A-Z]+\d+):([A-Z]+\d+)\)", sum_fn, formula, count=1, flags=re.I)
            if nxt != formula:
                formula = nxt
                continue
            nxt = re.sub(
                r"SUM\(([A-Z]+\d+)\)",
                lambda match: str(self.cell(current_sheet, int(re.match(r"[A-Z]+(\d+)", match.group(1)).group(1)), column_index_from_string(re.match(r"([A-Z]+)", match.group(1)).group(1)))),
                formula,
                count=1,
                flags=re.I,
            )
            if nxt != formula:
                formula = nxt
                continue
            break

        formula = re.sub(r"'([^']+)'!\$?([A-Z]+)\$?(\d+)", sheet_ref, formula)
        formula = re.sub(r"([A-Za-z][A-Za-z0-9 ]+)!\$?([A-Z]+)\$?(\d+)", sheet_ref, formula)

        def local_ref(match: re.Match) -> str:
            col = column_index_from_string(match.group(1))
            row = int(match.group(2))
            return str(self.cell(current_sheet, row, col))

        formula = re.sub(r"(?<![A-Za-z0-9_])\$?([A-Z]+)\$?(\d+)", local_ref, formula)

        formula = re.sub(r"\[.*?\][A-Za-z0-9 '$!]+", "0", formula)
        formula = re.sub(r"SUM\([^)]*\)", "0", formula, flags=re.I)
        formula = formula.replace("^", "**")
        if not re.fullmatch(r"[0-9eE.+*/() -]+", formula):
            return 0.0
        try:
            return float(eval(formula, {"__builtins__": {}}, {}))
        except Exception:
            return 0.0


SKIP_CLIENTS = {
    "actual",
    "acutal",
    "budget",
    "balance to cover",
    "no of days",
    "apc as per number of invoice",
    "total",
}


def parse_invoices(engine: WorkbookEngine) -> list[dict]:
    rows: list[dict] = []
    for key, label, sheet in MONTHS:
        ws = engine.sheet(sheet)
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 2).value
            if not name or not str(name).strip():
                continue
            if str(name).strip().lower() in SKIP_CLIENTS:
                continue
            invoice_no = ws.cell(r, 4).value
            if isinstance(invoice_no, str) and invoice_no.strip().startswith("="):
                continue
            if str(name).strip().startswith("=") or str(name).strip().isdigit():
                continue
            gst_raw = ws.cell(r, 3).value
            if isinstance(gst_raw, str) and gst_raw.strip().startswith("="):
                continue
            f_raw = ws.cell(r, 6).value
            if isinstance(f_raw, str) and f_raw.upper().startswith("=SUM("):
                continue
            rent = engine.cell(sheet, r, 6)
            packing = engine.cell(sheet, r, 7)
            transport = engine.cell(sheet, r, 8)
            mist = engine.cell(sheet, r, 9)
            discount = engine.cell(sheet, r, 10)
            breakage = engine.cell(sheet, r, 11)
            net = rent + packing + transport + mist - discount + breakage
            if (invoice_no in (None, "", 0, "0")) and abs(net) < 0.01:
                continue
            gst = round(net * 0.18, 2)
            gross = net + gst
            gstin = str(ws.cell(r, 3).value or "").strip()
            if gstin.startswith("="):
                gstin = ""
            date = str(ws.cell(r, 5).value or "").strip()
            if sheet == "April 26":
                deposit = 0.0
                received = engine.cell(sheet, r, 15)
                cash = 0.0
                next_month = engine.cell(sheet, r, 16)
                tds = 0.0
                pending_cell = engine.cell(sheet, r, 17)
                purchase = engine.cell(sheet, r, 19)
                pay_type = str(ws.cell(r, 18).value or "")
            elif sheet == "June 2026":
                deposit = engine.cell(sheet, r, 15)
                received = engine.cell(sheet, r, 16)
                tds = engine.cell(sheet, r, 17)
                cash = engine.cell(sheet, r, 18)
                next_month = 0.0
                pending_cell = engine.cell(sheet, r, 19)
                purchase = engine.cell(sheet, r, 21)
                pay_type = str(ws.cell(r, 20).value or "")
            else:
                deposit = engine.cell(sheet, r, 15)
                received = engine.cell(sheet, r, 16)
                cash = engine.cell(sheet, r, 17)
                tds = 0.0
                next_month = 0.0
                pending_cell = engine.cell(sheet, r, 18)
                purchase = engine.cell(sheet, r, 20)
                pay_type = str(ws.cell(r, 19).value or "")
            collected = received + cash + next_month
            if net < -1:
                status = "credit"
                pending = 0.0
            elif collected >= gross - 2:
                status = "collected_late" if next_month >= 2 and received + cash < 2 else "collected"
                pending = 0.0
            elif pending_cell >= 2:
                pending = pending_cell
                status = "partial" if collected >= 2 else "open"
            else:
                pending = max(0.0, round(gross - collected - tds, 2))
                if pending < 2:
                    pending = 0.0
                    status = "collected"
                else:
                    status = "partial" if collected >= 2 else "open"
            rows.append(
                {
                    "month": key,
                    "monthLabel": label,
                    "client": str(name).strip(),
                    "gstNumber": gstin,
                    "invoiceNo": str(invoice_no).strip() if invoice_no not in (None, 0, "0") else "",
                    "eventDate": date,
                    "rent": round(rent, 2),
                    "packing": round(packing, 2),
                    "transport": round(transport, 2),
                    "mist": round(mist, 2),
                    "discount": round(discount, 2),
                    "breakage": round(breakage, 2),
                    "net": round(net, 2),
                    "gst": round(gst, 2),
                    "gross": round(gross, 2),
                    "deposit": round(deposit, 2),
                    "received": round(received, 2),
                    "cash": round(cash, 2),
                    "nextMonth": round(next_month, 2),
                    "tds": round(tds, 2),
                    "collected": round(collected, 2),
                    "pending": round(pending, 2),
                    "status": status,
                    "paymentType": pay_type.strip(),
                    "purchaseValue": round(purchase, 2),
                    "roi": round(net / purchase, 4) if purchase else 0,
                }
            )
    return rows


def col_for_month(month_key: str) -> int:
    # B=April ... M=March
    order = [
        "2026-04",
        "2026-05",
        "2026-06",
        "2026-07",
        "2026-08",
        "2026-09",
        "2026-10",
        "2026-11",
        "2026-12",
        "2027-01",
        "2027-02",
        "2027-03",
    ]
    return order.index(month_key) + 2  # B=2


PNL_COST_ROWS = [
    (24, "fixed", "Rent with TDS"),
    (25, "fixed", "Salary"),
    (26, "fixed", "Electricity & water"),
    (27, "fixed", "Insurance"),
    (28, "fixed", "Petty cash"),
    (29, "fixed", "Internet"),
    (33, "ops", "OPS supply"),
    (34, "ops", "Transport"),
    (38, "admin", "CA, Fire, BMC, GST filing"),
    (39, "admin", "CCG breakage (unpaid)"),
    (40, "admin", "Business development"),
    (44, "nonctrl", "Bank charges"),
    (45, "nonctrl", "CAM / society"),
    (46, "nonctrl", "AMC"),
    (47, "nonctrl", "Interest on capex 8%"),
]


def parse_pnl_costs(engine: WorkbookEngine) -> dict:
    by_month = {}
    for i, key in enumerate(FY_MONTHS):
        col = i + 3  # C = April
        buckets = {"fixed": 0.0, "ops": 0.0, "admin": 0.0, "nonctrl": 0.0}
        lines = []
        for row, group, name in PNL_COST_ROWS:
            amt = round(engine.cell("Pnl", row, col), 2)
            buckets[group] += amt
            lines.append({"group": group, "name": name, "amount": amt})
        opex = buckets["fixed"] + buckets["ops"] + buckets["admin"] + buckets["nonctrl"]
        by_month[key] = {
            "fixed": round(buckets["fixed"], 2),
            "ops": round(buckets["ops"], 2),
            "admin": round(buckets["admin"], 2),
            "nonctrl": round(buckets["nonctrl"], 2),
            "total": round(opex, 2),
            "lines": lines,
        }
    return by_month


def parse_pending_sheet(engine: WorkbookEngine) -> list[dict]:
    sheet = "Pending Payments "
    rows = []
    for r in range(13, 22):
        name = engine.sheet(sheet).cell(r, 1).value
        if not name or str(name).strip().lower().startswith("total"):
            continue
        months = {}
        total = 0.0
        for i, key in enumerate(FY_MONTHS[:10]):
            amt = round(engine.cell(sheet, r, i + 2), 2)
            months[key] = amt
            total += amt
        rows.append({"client": str(name).strip(), "pending": round(total, 2), "byMonth": months})
    rows.sort(key=lambda x: -x["pending"])
    return rows


def main():
    revenue = WorkbookEngine(FOLDER / "Revenue Traker April 2026 to March 31st 2027 UPDATED.xlsx")
    pnl = WorkbookEngine(FOLDER / "Forecast PnL from April 2026 to March 2027.xlsx")
    invoices = parse_invoices(revenue)

    monthly = []
    for key, label, _ in MONTHS:
        subset = [row for row in invoices if row["month"] == key]
        rent = sum(r["rent"] for r in subset)
        packing = sum(r["packing"] for r in subset)
        transport = sum(r["transport"] for r in subset)
        mist = sum(r["mist"] for r in subset)
        discount = sum(r["discount"] for r in subset)
        breakage = sum(r["breakage"] for r in subset)
        net = rent + packing + transport + mist - discount + breakage
        gst = round(net * 0.18, 2)
        budget = round(ANNUAL_BUDGET * BUDGET_SHARE[key])
        purchase = sum(r["purchaseValue"] for r in subset)
        collected_same = sum(r["received"] + r["cash"] for r in subset)
        delayed = sum(r["nextMonth"] for r in subset)
        deposits = sum(r["deposit"] for r in subset)
        open_jobs = [r for r in subset if r["status"] in ("open", "partial")]
        collected_jobs = [r for r in subset if r["status"] in ("collected", "collected_late")]
        monthly.append(
            {
                "month": key,
                "label": label,
                "invoiceCount": len(subset),
                "rent": round(rent, 2),
                "packing": round(packing, 2),
                "transport": round(transport, 2),
                "mist": round(mist, 2),
                "discount": round(discount, 2),
                "breakage": round(breakage, 2),
                "net": round(net, 2),
                "gst": gst,
                "gross": round(net + gst, 2),
                "budget": budget,
                "vsBudget": round(net - budget, 2),
                "collectedSameMonth": round(collected_same, 2),
                "collectedLater": round(delayed, 2),
                "cashInOnBilling": round(collected_same + delayed, 2),
                "depositsHeld": round(deposits, 2),
                "received": round(collected_same, 2),
                "pending": round(sum(r["pending"] for r in subset), 2),
                "openCount": len(open_jobs),
                "collectedCount": len(collected_jobs),
                "collectionRate": round((collected_same + delayed) / (net + gst), 4) if net + gst else 0,
                "breakageRate": round(breakage / rent, 4) if rent else 0,
                "purchaseValue": round(purchase, 2),
                "roi": round(net / purchase, 4) if purchase else 0,
                "apc": round(net / len(subset), 2) if subset else 0,
            }
        )

    actual_keys = {m["month"] for m in monthly if m["invoiceCount"] > 0}
    monthly = [m for m in monthly if m["invoiceCount"] > 0]
    for i, m in enumerate(monthly):
        incoming_late = monthly[i - 1]["collectedLater"] if i else 0
        m["cashCollected"] = round(m["collectedSameMonth"] + incoming_late, 2)
        m["lateFromPriorMonth"] = round(incoming_late, 2)

    expenses_by_month = parse_pnl_costs(pnl)

    mix = {
        "rent": 0.86,
        "packing": 0.03,
        "transport": 0.01,
        "breakage": 0.10,
        "discount": 0.07,
    }
    actual_keys = {m["month"] for m in monthly}
    forecast_months = []
    for key in FY_MONTHS:
        budget = round(ANNUAL_BUDGET * BUDGET_SHARE[key])
        if key in actual_keys:
            continue
        rent = round(budget * mix["rent"], 2)
        packing = round(budget * mix["packing"], 2)
        transport = round(budget * mix["transport"], 2)
        breakage = round(budget * mix["breakage"], 2)
        discount = round(budget * mix["discount"], 2)
        net = round(rent + packing + transport + breakage - discount, 2)
        gst = round(net * 0.18, 2)
        forecast_months.append(
            {
                "month": key,
                "label": key,
                "invoiceCount": 0,
                "rent": rent,
                "packing": packing,
                "transport": transport,
                "mist": 0,
                "discount": discount,
                "breakage": breakage,
                "net": net,
                "gst": gst,
                "gross": round(net + gst, 2),
                "budget": budget,
                "vsBudget": 0,
                "received": 0,
                "pending": 0,
                "purchaseValue": 0,
                "roi": 0,
                "apc": 0,
                "forecast": True,
            }
        )

    pnl_months = []
    for item in monthly + forecast_months:
        key = item["month"]
        exp = expenses_by_month.get(
            key, {"fixed": 0, "ops": 0, "admin": 0, "nonctrl": 0, "total": 0, "lines": []}
        )
        ebitda = item["net"] - exp["total"]
        labels = {
            "2026-10": "Oct 2026",
            "2026-11": "Nov 2026",
            "2026-12": "Dec 2026",
            "2027-01": "Jan 2027",
            "2027-02": "Feb 2027",
            "2027-03": "Mar 2027",
        }
        pnl_months.append(
            {
                **item,
                "label": item.get("label") if "Apr" in str(item.get("label")) or "May" in str(item.get("label")) or "Jun" in str(item.get("label")) or "Jul" in str(item.get("label")) or "Aug" in str(item.get("label")) or "Sep" in str(item.get("label")) else labels.get(key, item["label"]),
                "fixedCost": exp["fixed"],
                "opsCost": exp["ops"],
                "adminCost": exp["admin"],
                "nonctrlCost": exp.get("nonctrl", 0),
                "totalOpex": exp["total"],
                "ebitda": round(ebitda, 2),
                "ebitdaMargin": round(ebitda / item["net"], 4) if item["net"] else 0,
                "forecast": bool(item.get("forecast")),
            }
        )

    cash = {}
    for key in ["2026-04", "2026-05", "2026-06", "2026-07", "2026-08"]:
        col = col_for_month(key)
        cash[key] = {
            "bank": round(pnl.cell("Cash flow ", 18, col), 2),
            "cash": round(pnl.cell("Cash flow ", 19, col), 2),
        }

    pending_clients = {}
    for row in invoices:
        if row["pending"] <= 1:
            continue
        pending_clients.setdefault(row["client"], 0)
        pending_clients[row["client"]] += row["pending"]

    capex = {
        "samirShare": 0.74,
        "karanShare": 0.26,
        "samirTarget": 4_810_000,
        "karanTarget": 1_690_000,
        "samirInvested": 4_745_895 + 64_105,
        "karanInvested": 1_210_000 + 480_000,
        "ccgVendors": [
            {"name": "Harbour Artevo, Selite, Rona", "actual": 935000},
            {"name": "Thanor Pottery", "actual": 146319 + 313666},
            {"name": "Viraj Nirman", "actual": 938926},
            {"name": "SK Manekia", "actual": 204016 + 202407},
            {"name": "Neeti Udyog", "actual": 326683},
        ],
    }
    capex["totalTarget"] = capex["samirTarget"] + capex["karanTarget"]
    capex["totalInvested"] = capex["samirInvested"] + capex["karanInvested"]
    capex["balance"] = capex["totalTarget"] - capex["totalInvested"]

    actual_months = [m for m in pnl_months if not m.get("forecast")]
    billed_net = sum(m["net"] for m in actual_months)
    gst_pass = sum(m["gst"] for m in actual_months)
    deposits = sum(m.get("depositsHeld") or 0 for m in actual_months)
    cash_in = sum(m.get("cashCollected") or 0 for m in actual_months)
    still_owed = sum(m.get("pending") or 0 for m in actual_months)
    breakage = sum(m["breakage"] for m in actual_months)
    rent = sum(m["rent"] for m in actual_months)
    open_invoices = [r for r in invoices if r["status"] in ("open", "partial")]
    collected_invoices = [r for r in invoices if r["status"] in ("collected", "collected_late")]
    late_invoices = [r for r in invoices if r["status"] == "collected_late"]
    story = {
        "billedNet": round(billed_net, 2),
        "gstPassThrough": round(gst_pass, 2),
        "depositsNotRevenue": round(deposits, 2),
        "cashCollected": round(cash_in, 2),
        "stillOwed": round(still_owed, 2),
        "breakage": round(breakage, 2),
        "rent": round(rent, 2),
        "breakageOfRent": round(breakage / rent, 4) if rent else 0,
        "openJobs": len(open_invoices),
        "collectedJobs": len(collected_invoices),
        "paidLateJobs": len(late_invoices),
        "collectionRate": round(cash_in / (billed_net + gst_pass), 4) if billed_net + gst_pass else 0,
    }

    payload = {
        "source": "Switch Rentals live workbooks FY 2026-27",
        "company": {
            "legalName": "Switch Rental Services LLP",
            "gstin": "27AFHFS2025K1ZV",
            "hsn": "997323",
            "address": "Ground floor, Gupta Mill Estate, Devidayal Mill Compound, Magazine Street, Darukhana, Mumbai 400010",
            "phones": ["9125660485", "9820167690"],
            "emails": ["info@switchrental.in", "samir@switchrental.in"],
            "bank": {
                "name": "ICICI, Colaba Branch",
                "accountName": "Switch Rental Services LLP",
                "accountNo": "104305003604",
                "ifsc": "ICIC0001043",
            },
            "packingRate": 0.03,
            "gstRate": 0.18,
            "lateReturn": "25% extra till 6pm, 100% extra per extra day after 2pm next day",
            "toteBoxCharge": 1850,
            "invoiceTerms": [
                "Rates are per piece per day.",
                "Chalan is issued on delivery. Return the next day between 11am and 2pm.",
                "Late return: 25% extra till 6pm, then 100% of rental per extra day.",
                "Lost tote box: INR 1,850.",
                "Rinse items before return.",
                "Packing is 3% of rental. GST 18% on rent and breakage (CGST 9% + SGST 9%). HSN 997323.",
            ],
        },
        "budget": {
            "annualNet": ANNUAL_BUDGET,
            "shares": BUDGET_SHARE,
            "forwardMix": {"rent": 0.86, "packing": 0.03, "transport": 0.01, "breakage": 0.10, "discount": 0.07},
        },
        "story": story,
        "invoices": invoices,
        "monthly": pnl_months,
        "expensesByMonth": expenses_by_month,
        "cash": cash,
        "pendingTracker": parse_pending_sheet(revenue),
        "pendingByClient": [
            {"client": k, "pending": round(v, 2)} for k, v in sorted(pending_clients.items(), key=lambda x: -x[1])
        ],
        "capex": capex,
        "partners": [
            {"name": "Samir Chhabria", "share": 0.74, "role": "Operating partner"},
            {"name": "Karan Khiani", "share": 0.26, "role": "Partner"},
        ],
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2))
    print(f"Wrote {OUT} invoices={len(invoices)} months={len(pnl_months)}")
    for m in pnl_months:
        print(
            f"{m['label']}: inv={m['invoiceCount']} net={m['net']:.0f} budget={m['budget']} ebitda={m['ebitda']:.0f} pending={m['pending']:.0f}"
        )


if __name__ == "__main__":
    main()
